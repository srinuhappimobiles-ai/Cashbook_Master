import io
import os
import re
import math
import subprocess
import platform
import numpy as np
import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Happi Cashbook Master - Excel Automation",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
    <style>
    .block-container { 
        padding-top: 1.2rem !important; 
        padding-bottom: 1rem !important; 
        padding-left: 1.5rem !important; 
        padding-right: 1.5rem !important; 
        max-width: 100% !important;
    }
    #MainMenu {visibility: hidden !important;}
    footer {visibility: hidden !important;}
    .main-card {
        background: #ffffff;
        border: 1px solid #e2e8f0;
        border-radius: 8px;
        padding: 24px;
        margin-bottom: 20px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05);
    }
    .main-title { 
        font-size: 22px; 
        font-weight: 700; 
        color: #107c41; 
        margin-bottom: 6px; 
        display: flex; 
        align-items: center; 
        gap: 10px; 
    }
    .stSidebar { background-color: #f8fafc; }
    </style>
""", unsafe_allow_html=True)

DEFAULT_BRANCHES = [
    {"CODE": "ADBD", "BRANCH": "ADILABAD"}, {"CODE": "AMP", "BRANCH": "AMALAPURAM"},
    {"CODE": "AMPT", "BRANCH": "AMEERPET"}, {"CODE": "ANTP", "BRANCH": "ANANTAPUR"},
    {"CODE": "ARMU", "BRANCH": "ARMOOR"}, {"CODE": "ATMKR", "BRANCH": "ATMAKUR"},
    {"CODE": "BDHN", "BRANCH": "BODHAN"}, {"CODE": "BDPL", "BRANCH": "BODUPPAL"},
    {"CODE": "BG", "BRANCH": "BHUVANAGIRI"}, {"CODE": "BVRM", "BRANCH": "BHIMAVARAM"},
    {"CODE": "CHND", "BRANCH": "CHANDANAGAR"}, {"CODE": "CHNT", "BRANCH": "CHINTAL"},
    {"CODE": "DBGS", "BRANCH": "DABAGARDENS"}, {"CODE": "DBGS-2", "BRANCH": "DABAGARDENS-2"},
    {"CODE": "DVKD", "BRANCH": "DEVARAKONDA"}, {"CODE": "DVRM", "BRANCH": "DHARMAVRAM"},
    {"CODE": "ECIL", "BRANCH": "ECIL"}, {"CODE": "ELR", "BRANCH": "ELURU"},
    {"CODE": "GDVK", "BRANCH": "GODHAVARIKHANI"}, {"CODE": "GJWK", "BRANCH": "GAJUWAKA"},
    {"CODE": "GJWL", "BRANCH": "GAJWEL"}, {"CODE": "GNT", "BRANCH": "GUNTUR"},
    {"CODE": "GNT2", "BRANCH": "GUNTUR2"}, {"CODE": "GTKL", "BRANCH": "GUNTAKAL"},
    {"CODE": "GWD", "BRANCH": "GADWAL"}, {"CODE": "HAL", "BRANCH": "HALIYA"},
    {"CODE": "HNMK", "BRANCH": "HANUMAKONDA"}, {"CODE": "HUP", "BRANCH": "HINDUPUR"},
    {"CODE": "JCL", "BRANCH": "JADCHERLA"}, {"CODE": "JNGN", "BRANCH": "JANGAON"},
    {"CODE": "JTL", "BRANCH": "JAGTIAL"}, {"CODE": "KDGM", "BRANCH": "KALYANADURGAM"},
    {"CODE": "KDR", "BRANCH": "KADIRI"}, {"CODE": "KDR2", "BRANCH": "KADIRI-2"},
    {"CODE": "KKP", "BRANCH": "KUKATPALLY"}, {"CODE": "KMGH", "BRANCH": "KHARMANGHAT"},
    {"CODE": "KMM", "BRANCH": "KHAMMAM"}, {"CODE": "KMM2", "BRANCH": "KHAMMAM 2"},
    {"CODE": "KPM", "BRANCH": "KUPPAM"}, {"CODE": "KRKH", "BRANCH": "KHARKHANA"},
    {"CODE": "KRLA", "BRANCH": "KORUTLA"}, {"CODE": "KRMN", "BRANCH": "KARIMNAGAR"},
    {"CODE": "KRNL", "BRANCH": "KURNOOL"}, {"CODE": "KRNL2", "BRANCH": "KURNOOL2"},
    {"CODE": "KZP", "BRANCH": "KAZIPET"}, {"CODE": "MCI", "BRANCH": "MANCHERIAL"},
    {"CODE": "MDPL", "BRANCH": "MADANAPALLI"}, {"CODE": "MDPR", "BRANCH": "MADHAPUR"},
    {"CODE": "MDPT", "BRANCH": "MANDAPETA"}, {"CODE": "MHBR", "BRANCH": "MAHABUBNAGAR"},
    {"CODE": "MLKJ", "BRANCH": "MALKAJGIRI"}, {"CODE": "MRGA", "BRANCH": "MIRYALAGUDA"},
    {"CODE": "MTM", "BRANCH": "MACHILIPATNAM"}, {"CODE": "MVP", "BRANCH": "MVP COLONY"},
    {"CODE": "NDD", "BRANCH": "NIDADAVOLE"}, {"CODE": "NDL", "BRANCH": "NANDYALA"},
    {"CODE": "NGKL", "BRANCH": "NAGARKURNOOL"}, {"CODE": "NKRL", "BRANCH": "NAKREKAL"},
    {"CODE": "NLG", "BRANCH": "NALGONDA"}, {"CODE": "NRKD", "BRANCH": "NARAYANKHED"},
    {"CODE": "NRML", "BRANCH": "NIRMAL"}, {"CODE": "NRSP", "BRANCH": "NARASANNAPETA"},
    {"CODE": "NSMP", "BRANCH": "NARSAMPET"}, {"CODE": "NSPT", "BRANCH": "NARSIPATNAM"},
    {"CODE": "NZVD", "BRANCH": "NUZVID"}, {"CODE": "ONG", "BRANCH": "ONGOLE"},
    {"CODE": "PDPL", "BRANCH": "PEDDAPALLI"}, {"CODE": "PDPM", "BRANCH": "PEDDAPURAM"},
    {"CODE": "PIL", "BRANCH": "PILERU"}, {"CODE": "PLM", "BRANCH": "PALAMANER"},
    {"CODE": "PSA", "BRANCH": "PALASA"}, {"CODE": "PVA", "BRANCH": "PALAVANCHA"},
    {"CODE": "RCT", "BRANCH": "RAYACHOTI"}, {"CODE": "RJY", "BRANCH": "RAJAMUNDRY"},
    {"CODE": "RMTP", "BRANCH": "RAMANTHAPUR"}, {"CODE": "RTCX", "BRANCH": "RTC X ROAD"},
    {"CODE": "SDPT", "BRANCH": "SIDDIPET"}, {"CODE": "SDR", "BRANCH": "S.D.ROAD"},
    {"CODE": "SHDR", "BRANCH": "SHADNAGAR"}, {"CODE": "SHPR", "BRANCH": "SHAPUR"},
    {"CODE": "SKKM", "BRANCH": "SRIKAKULAM"}, {"CODE": "SMBD", "BRANCH": "SHAMSHABAD"},
    {"CODE": "SNGR", "BRANCH": "SANGAREDDY"}, {"CODE": "SPT", "BRANCH": "SOMPETA"},
    {"CODE": "SRN", "BRANCH": "S.R.NAGAR"}, {"CODE": "SRNR", "BRANCH": "SAROORNAGAR"},
    {"CODE": "SRPT", "BRANCH": "SURYAPET"}, {"CODE": "STNR", "BRANCH": "SANTOSHNAGAR"},
    {"CODE": "TDPG", "BRANCH": "TADEPALLIGUDEM"}, {"CODE": "TDPT", "BRANCH": "TADIPATRI"},
    {"CODE": "TDU", "BRANCH": "TANDUR"}, {"CODE": "TEK", "BRANCH": "TEKKALI"},
    {"CODE": "TN", "BRANCH": "TUNI"}, {"CODE": "TNK", "BRANCH": "TANUKU"},
    {"CODE": "TNL", "BRANCH": "TENALI"}, {"CODE": "TPT", "BRANCH": "TIRUPATHI"},
    {"CODE": "TPT2", "BRANCH": "TIRUPATHI 2"}, {"CODE": "UPL", "BRANCH": "UPPAL"},
    {"CODE": "VIJ-1", "BRANCH": "VIJAYAWADA 1"}, {"CODE": "VIJ-3", "BRANCH": "VIJAYAWADA 3"},
    {"CODE": "VIJ-4", "BRANCH": "VIJAYAWADA 4"}, {"CODE": "VNSP", "BRANCH": "VANASTALIPURAM"},
    {"CODE": "VZM", "BRANCH": "VIZIANAGARAM"}, {"CODE": "VZM2", "BRANCH": "VIZIANAGARAM 2"},
    {"CODE": "WGL", "BRANCH": "WARANGAL"}, {"CODE": "WGL2", "BRANCH": "WARANGAL 2"},
    {"CODE": "ZB", "BRANCH": "ZAHEERABAD"}
]

HEADERS = [
    "SL.No.", "CODE", "BRANCH", "OPENING BALANCE", "DEPOSIT", "DENOMINATION", 
    "AddinGS", "PENDING APPRVLS", "FINANCE AMNT", "SR", "SWEEPER SALARY", 
    "EDITS", "APX SHORTAGE", "(KSP)'Sir's Approvals", "CLOSING BALANCE", "REMARKS"
]

SUPPORTED_TYPES = ["xlsx", "xls", "xlsm", "xlsb", "csv"]

def normalize_key(s):
    if not s: return ""
    return re.sub(r"[^A-Za-z0-9]", "", str(s)).upper()

def format_excel_formula(num_list):
    if not num_list: return ""
    clean_ints = [str(int(x)) for x in num_list if x > 0]
    if not clean_ints: return ""
    if len(clean_ints) == 1: return clean_ints[0]
    return "=" + "+".join(clean_ints)

# Session state initialization
if "dr_balances" not in st.session_state:
    st.session_state.dr_balances = {}
if "addins_data" not in st.session_state:
    st.session_state.addins_data = {}

# --- SIDEBAR CONTROL PANEL ---
with st.sidebar:
    st.markdown("### 🏢 Happi Control Hub")
    
    st.markdown("#### 👥 Work Assignment Mode")
    work_mode = st.radio(
        "Select Mode:",
        ["👤 Single Cashier (All Stores)", "👥 Two Cashiers (Split 50-50)"],
        label_visibility="collapsed"
    )

    all_branches = sorted(DEFAULT_BRANCHES, key=lambda x: str(x.get("BRANCH", "")).upper())
    selected_branches = all_branches
    sheet_title = "MASTER CASHBOOK"

    if work_mode == "👥 Two Cashiers (Split 50-50)":
        mid_point = math.ceil(len(all_branches) / 2)
        c1_branches = all_branches[:mid_point]
        c2_branches = all_branches[mid_point:]
        cashier_view = st.selectbox("Current Cashier Sheet:", [f"Cashier 1 ({len(c1_branches)} Stores)", f"Cashier 2 ({len(c2_branches)} Stores)"])
        if "Cashier 1" in cashier_view:
            selected_branches = c1_branches
            sheet_title = "CASHIER 1 CASHBOOK"
        else:
            selected_branches = c2_branches
            sheet_title = "CASHIER 2 CASHBOOK"

    st.markdown("---")
    st.markdown("#### 📥 Direct File Ingestion")
    
    with st.expander("📂 1. Apex Dr Balance File", expanded=True):
        ho_dump_file = st.file_uploader("Upload Apex Dr Balance", type=SUPPORTED_TYPES, key="ho_dump")

    with st.expander("📥 2. Addins Dump File", expanded=True):
        addins_dump_file = st.file_uploader("Upload Addins Dump", type=SUPPORTED_TYPES, key="addins_dump")

    st.markdown("---")
    if st.button("🧹 Clear Ingested Data", use_container_width=True):
        st.session_state.dr_balances = {}
        st.session_state.addins_data = {}
        st.rerun()

# 1. Process Apex Dump
if ho_dump_file:
    try:
        df_dump = pd.read_csv(ho_dump_file) if ho_dump_file.name.endswith(".csv") else pd.read_excel(ho_dump_file)
        b_col, a_col = df_dump.columns[0], df_dump.columns[1] if len(df_dump.columns) > 1 else df_dump.columns[0]
        for col in df_dump.columns:
            if "branch" in str(col).lower() or "store" in str(col).lower(): b_col = col
            if "balance" in str(col).lower() or "opening" in str(col).lower() or "amount" in str(col).lower(): a_col = col

        dump_dict = {}
        for _, row in df_dump.iterrows():
            b_val, raw_amt = normalize_key(row[b_col]), str(row[a_col]).replace(",", "").strip()
            match = re.search(r"(\d+\.?\d*)", raw_amt)
            if match and b_val:
                dump_dict[b_val] = int(round(float(match.group(1))))

        st.session_state.dr_balances = dump_dict
        st.sidebar.success(f"✅ Loaded {len(dump_dict)} Opening Balances!")
    except Exception as e:
        st.sidebar.error(f"Apex Error: {e}")

# 2. Process Addins Dump
if addins_dump_file:
    try:
        df_addins = pd.read_csv(addins_dump_file) if addins_dump_file.name.endswith(".csv") else pd.read_excel(addins_dump_file)
        b_col, a_col = df_addins.columns[0], df_addins.columns[1] if len(df_addins.columns) > 1 else df_addins.columns[0]
        for col in df_addins.columns:
            if "branch" in str(col).lower() or "store" in str(col).lower(): b_col = col
            if "amount" in str(col).lower() or "total" in str(col).lower() or "addin" in str(col).lower(): a_col = col

        addins_dict = {}
        for _, row in df_addins.iterrows():
            b_val, raw_amt = normalize_key(row[b_col]), str(row[a_col]).replace(",", "").strip()
            match = re.search(r"(\d+\.?\d*)", raw_amt)
            if match and b_val:
                clean_amt = int(round(float(match.group(1))))
                if clean_amt > 0:
                    addins_dict.setdefault(b_val, []).append(clean_amt)

        st.session_state.addins_data = addins_dict
        st.sidebar.success(f"✅ Loaded {len(addins_dict)} Store Addins!")
    except Exception as e:
        st.sidebar.error(f"Addins Error: {e}")

# --- BUILD EXCEL 2010 WORKBOOK IN MEMORY ---
def generate_excel_2010_file(target_branches, title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    # Styles
    header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    bold_font = Font(name="Calibri", size=11, bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='A0A0A0'),
        right=Side(style='thin', color='A0A0A0'),
        top=Side(style='thin', color='A0A0A0'),
        bottom=Side(style='thin', color='A0A0A0')
    )

    # 1. Write Headers (Row 1)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[1].height = 26

    # 2. Write Store Rows
    dr_map = st.session_state.dr_balances
    addins_map = st.session_state.addins_data

    for row_idx, b in enumerate(target_branches, start=2):
        b_name = b.get("BRANCH", "")
        b_code = b.get("CODE", "")
        norm_name = normalize_key(b_name)
        norm_code = normalize_key(b_code)

        # Opening Balance
        op_val = dr_map.get(norm_name, dr_map.get(norm_code, ""))

        # Addins Formula
        vouchers = addins_map.get(norm_name, addins_map.get(norm_code, []))
        addin_val = format_excel_formula(vouchers) if vouchers else ""

        # Closing Balance Formula: =D2-SUM(E2:N2)
        closing_formula = f"=D{row_idx}-SUM(E{row_idx}:N{row_idx})"

        row_values = [
            row_idx - 1,      # A: SL.No.
            b_code,           # B: CODE
            b_name,           # C: BRANCH
            op_val,           # D: OPENING BALANCE
            "",               # E: DEPOSIT
            "",               # F: DENOMINATION
            addin_val,        # G: AddinGS
            "",               # H: PENDING APPRVLS
            "",               # I: FINANCE AMNT
            "",               # J: SR
            "",               # K: SWEEPER SALARY
            "",               # L: EDITS
            "",               # M: APX SHORTAGE
            "",               # N: (KSP)'Sir's Approvals
            closing_formula,  # O: CLOSING BALANCE
            ""                # P: REMARKS
        ]

        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if str(val).startswith("="):
                cell.value = str(val)
            elif isinstance(val, (int, float)):
                cell.value = val
                cell.number_format = '#,##0'
            else:
                cell.value = val

            cell.font = bold_font if col_idx in [1, 2, 3, 15] else data_font
            cell.alignment = Alignment(horizontal="center" if col_idx in [1, 2] else "left" if col_idx in [3, 16] else "right", vertical="center")
            cell.border = thin_border

        ws.row_dimensions[row_idx].height = 20

    # Auto-adjust column widths
    column_widths = {
        1: 8, 2: 12, 3: 24, 4: 20, 5: 14, 6: 16, 7: 16, 8: 18,
        9: 16, 10: 14, 11: 18, 12: 14, 13: 16, 14: 22, 15: 20, 16: 22
    }
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# Generate Excel Binary
excel_data = generate_excel_2010_file(selected_branches, sheet_title)
file_name = f"{sheet_title.replace(' ', '_')}_2010.xlsx"

# Function to launch directly into desktop MS Excel
def open_in_desktop_excel(data, fname):
    local_path = os.path.abspath(fname)
    with open(local_path, "wb") as f:
        f.write(data)
    
    current_os = platform.system()
    if current_os == "Windows":
        os.startfile(local_path)
    elif current_os == "Darwin":
        subprocess.call(["open", local_path])
    else:
        subprocess.call(["xdg-open", local_path])

# --- MAIN PAGE DISPLAY ---
st.markdown(f'<div class="main-title">📊 HAPPI MOBILES - DESKTOP EXCEL 2010 WORKSPACE</div>', unsafe_allow_html=True)

st.markdown(f"""
<div class="main-card">
    <h4 style="margin-top: 0; color: #1e293b;">⚡ Master Sheet Ready ({len(selected_branches)} Stores Active)</h4>
    <p style="color: #64748b; font-size: 14px; margin-bottom: 20px;">
        All Apex Dr Balances, Addins formulas, and Closing Balance calculations (<code>=D2-SUM(E2:N2)</code>) have been processed.
        Click below to launch directly into your native <b>Microsoft Excel 2010</b> software with full keyboard shortcut support (<code>Ctrl+D</code>, <code>Alt+HBA</code>, <code>Alt+=</code>).
    </p>
</div>
""", unsafe_allow_html=True)

col_act1, col_act2 = st.columns(2)

with col_act1:
    if st.button("🚀 Open Directly in Desktop Excel 2010", type="primary", use_container_width=True):
        try:
            open_in_desktop_excel(excel_data, file_name)
            st.success(f"✅ Launched {file_name} in Microsoft Excel 2010!")
        except Exception as e:
            st.info("💡 Running on Cloud: Please use the Download button on the right to open directly in your Excel 2010.")

with col_act2:
    st.download_button(
        label="📥 Download & Open Excel 2010 File",
        data=excel_data,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

st.markdown("---")
st.markdown("#### 📋 Live Preview of Mapped Data")

preview_rows = []
dr_map = st.session_state.dr_balances
addins_map = st.session_state.addins_data

for idx, b in enumerate(selected_branches, start=1):
    b_name = b.get("BRANCH", "")
    b_code = b.get("CODE", "")
    norm_name = normalize_key(b_name)
    norm_code = normalize_key(b_code)
    
    op_val = dr_map.get(norm_name, dr_map.get(norm_code, ""))
    vouchers = addins_map.get(norm_name, addins_map.get(norm_code, []))
    addin_val = format_excel_formula(vouchers) if vouchers else ""
    
    preview_rows.append({
        "SL.No.": idx,
        "CODE": b_code,
        "BRANCH": b_name,
        "OPENING BALANCE": op_val,
        "AddinGS": addin_val,
        "CLOSING BALANCE FORMULA": f"=D{idx+1}-SUM(E{idx+1}:N{idx+1})"
    })

st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, height=450)
